
import openpyxl


def getPurchasePrice(sheet_obj, row):
    val = float(sheet_obj.cell(row, 51).value) * \
        (100 - float(sheet_obj.cell(row, 50).value)) / 10000
    val = val * (100 + float(sheet_obj.cell(row, 82).value) +
                 float(sheet_obj.cell(row, 84).value))
    val = round(val, 4)
    return val

# (in d-monthname-yy format) InvDate[5] or InvDay[6] + enum[InvMonth][7] + InvYear(only tens and units)[8]


def getPurchaseDate(sheet_obj, row):
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    val = sheet_obj.cell(
        row, 7).value + '-' + months[int(sheet_obj.cell(row, 8).value)-1] + '-' + str(int(sheet_obj.cell(row, 9).value))[2:]
    return val

# (in dd/mm/yyyy format)ExpDate[41]   (replace - with /) or expDay[42] + Expmonth[43] + Expyear[44]


def getExpiryDate(sheet_obj, row):
    day = sheet_obj.cell(row, 43).value.replace(" ", "")
    month = sheet_obj.cell(row, 44).value.replace(" ", "")
    year = sheet_obj.cell(row, 45).value.replace(" ", "")
    if(len(sheet_obj.cell(row, 44).value) == 1):
        month = '0' + month
    val = day + '/' + month + '/' + year
    return val


def readModifiedRow(sheet_obj, row, colList):
    new_val = [0]*len(colList)
    for i in range(len(colList)):
        if colList[i] != 100:
            new_val[i] = sheet_obj.cell(row, colList[i]+1).value
        else:
            new_val[i] = ' '
    try:
        new_val[8] = float(sheet_obj.cell(row, 82).value) + \
            float(sheet_obj.cell(row, 84).value)
        new_val[7] = getPurchasePrice(sheet_obj, row)
        new_val[12] = getPurchaseDate(sheet_obj, row)
        new_val[13] = getExpiryDate(sheet_obj, row)
        new_val[22] = (float(new_val[10]) - new_val[7]) / \
            float(new_val[10])  # ((MRP-Purchase.Price)/MRP)
        # (MRP - PurchasePrice)
        new_val[23] = float(new_val[10]) - new_val[7]
        new_val[22] = str(new_val[22])[:7]
        new_val[23] = str(new_val[23])[:7]
        new_val[5] = int(float(new_val[5]))
        new_val[6] = int(float(new_val[6]))
    except Exception as e:
        raise Exception('Check no. of rows')
    return new_val


def printRows(rows):
    headings = ['Invoice number', 'Code\t', 'Name\t', 'Type\t', 'Packing Number', 'Quantity', 'Free\t', 'PurchacePrice',
                'GST\t', 'Discount', 'MRP\t', 'BatchNo\t', 'PurchaseDate', 'ExpDate\t', 'Vendor Name \t', 'Manufacturer Name',
                'Compostitions\t', 'Rack No\t\t', 'HsnCode\t', 'Schedule Type\t', 'GST includedIn Rt', 'department\t', 'Margin %\t',
                'Margin\t\t']
    mainHeadings = [0, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 18, 22]
    print("TABLE:\n")
    # if(len(rows)):
    for cellNo in mainHeadings:
        print(headings[cellNo].replace("\t", ""), end=" | ")
    print('\n')
    for row in rows:
        for cellNo in mainHeadings:
            if(row[cellNo] == ' '):
                print('blank', end=" | ")
            else:
                print(row[cellNo], end=" | ")
        print('\n')
    # else:
    #     for i in range(len(headings)):
    #         print(headings[i], end="\t-\t")
    #         for j in range(len(rows)):
    #             if(rows[j][i] == ' '):
    #                 print("blank", end="\t|\t")
    #             else:
    #                 print(rows[j][i], end="\t|\t")
    #         print("\n")
    return


def displayUpdates(modified_rows):
    headings = ["Name", "HSN Code", "Batch#", "Unit Size", "Items/packs", "Total packs", "Free quantity", "MRP", "Rate", "Disc",
                "Disc", "GST", "Exp Date", "Purch Date"]
    # To do


def updateRows(main_wb, modified_rows, path):
    main_sheet = main_wb.active
    cur_line = main_sheet.max_row + 1
    for row in modified_rows:
        for i in range(len(row)):
            main_sheet.cell(cur_line, i+1).value = row[i]
        cur_line += 1
    main_wb.save(path)
    return


# Update rows from sheet_obj to main_sheet
def updateSheet(main_wb, bill_wb, colList, path):
    sheet_obj = bill_wb.active
    modified_rows = []
    print("Readings " + str(sheet_obj.max_row - 1) + ' rows...')
    for row in range(2, sheet_obj.max_row + 1):
        modified_row = readModifiedRow(sheet_obj, row, colList)
        modified_rows.append(modified_row)
    printRows(modified_rows)
    print("Continue with this?\n")
    opt = input("Y/N ")
    if(opt == 'Y'):
        print("Updating...")
        updateRows(main_wb, modified_rows, path)
    else:
        print("Aborting...")
    return


def openSheet(path):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    return sheet_obj


def testUpdate(sheet_wb, path):
    sheet_obj = sheet_wb.active
    print(sheet_obj.max_row)
    sheet_wb.save(path)
    print(sheet_obj.max_row)


# path_main = "./trial-main.xlsx"
cur_main_path = "./Current_Main/main.xlsx"
bill_path = "./bills/new_bill.xlsx"
path_test = "test.csv"
path_test_xlsx = "./bill.xlsx"
updated_main_path = "./upd/main.xlsx"
path_updated_bill = "./upd/bill.xlsx"

if __name__ == '__main__':
    print("Hello user")
    colList = [4, 9, 37, 39, 39, 45, 46, 100, 100, 100, 58, 40,
               100, 100, 100, 65, 100, 100, 80, 100, 100, 100, 100, 100]
    main_wb = openpyxl.load_workbook(cur_main_path)
    new_bill_wb = openpyxl.load_workbook(bill_path)
    updateSheet(main_wb, new_bill_wb, colList, updated_main_path)
